"""
API Server — BizAgent HTTP 服务（含任务追踪 + 会话持久化）
==========================================================
运行: python api_server.py
访问: http://localhost:5001/docs (Swagger UI)
     http://localhost:5000/api/health

对话端點:
  POST /chat                              — 统一对话接口（smart / analysis）

任务追踪端點:
  GET  /api/tasks/{session_id}            — 列出会话的任务
  GET  /api/tasks/{session_id}/{task_id}  — 获取任务完整执行日志
  GET  /api/logs                          — 读取最近的结构化日志
  GET  /api/stats                         — 系统统计信息

会话管理端點:
  GET  /api/sessions                      — 查看活跃会话列表
  GET  /api/sessions/{session_id}/messages — 获取会话消息历史
  DELETE /api/sessions/{session_id}       — 删除会话及其数据

文件支持:
  .txt / .csv / .json / .md / .pdf / .docx / .xlsx
  .png / .jpg / .jpeg（仅记录文件名）

启用持久化: POST /chat 传入 enable_persistence=true
"""

import os
import io
import re
import json
import uuid
import base64
import logging
import sys

# 确保能找到 src/ 下的包
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from typing import Optional, Literal
from pydantic import BaseModel, Field
import uvicorn

from minimal_agent import BizAgent, session_store
from minimal_agent.task_tracker import TaskTracker
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# ============================================================
# FastAPI 应用
# ============================================================
app = FastAPI(
    title="BizAgent API — 商务智能助手",
    description="基于 DeepSeek + 三层 Prompt + Function Calling 的商务智能 Agent HTTP 服务",
    version="1.0.0",
)

# CORS — 允许跨域请求
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 挂载图表静态目录
_charts_dir = os.path.join(os.path.dirname(__file__), "charts")
os.makedirs(_charts_dir, exist_ok=True)
app.mount("/charts", StaticFiles(directory=_charts_dir), name="charts")

# ============================================================
# 会话管理
# ============================================================
sessions: dict[str, BizAgent] = {}

API_KEY = os.getenv("DEEPSEEK_API_KEY")
BASE_URL = os.getenv("DEEPSEEK_BASE_URL")

if not API_KEY:
    raise RuntimeError("DEEPSEEK_API_KEY 未配置，请检查 .env 文件")


def get_or_create_agent(session_id: str = None, enable_persistence: bool = False) -> tuple[str, BizAgent]:
    """获取或创建会话对应的 BizAgent 实例"""
    if not session_id:
        session_id = uuid.uuid4().hex[:12]
    if session_id not in sessions:
        agent = BizAgent(
            API_KEY, BASE_URL,
            session_id=session_id,
            enable_persistence=enable_persistence,
        )
        sessions[session_id] = agent
        logger.info("[Session] 创建新会话: %s (persist=%s)", session_id, enable_persistence)
    return session_id, sessions[session_id]


# ============================================================
# 文件解析
# ============================================================

SUPPORTED_EXTENSIONS = {
    ".txt", ".csv", ".json", ".md", ".log", ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg",
    ".pdf", ".docx", ".xlsx", ".xls",
}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"}


def _parse_text(content: bytes) -> str:
    """以 UTF-8 解码文本，失败时尝试 GBK 回退"""
    for enc in ["utf-8", "gbk", "gb2312", "latin-1"]:
        try:
            return content.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="replace")


def _parse_pdf(content: bytes) -> str:
    """用 PyMuPDF 提取 PDF 文字"""
    import fitz
    doc = fitz.open(stream=content, filetype="pdf")
    texts = []
    for page in doc:
        texts.append(page.get_text())
    doc.close()
    return "\n".join(texts)


def _parse_docx(content: bytes) -> str:
    """用 python-docx 提取 Word 文字"""
    import docx
    doc = docx.Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs)


def _parse_xlsx(content: bytes) -> str:
    """用 openpyxl 提取 Excel 文字"""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            line = "\t".join(cells)
            if line.strip():
                lines.append(line)
    wb.close()
    return "\n".join(lines)


def parse_file_content(file_content: str, file_name: str) -> str:
    """
    解析上传的文件内容。

    Args:
        file_content: Base64 编码的文件数据
        file_name: 原文件名（用于判断类型）

    Returns:
        提取出的文本内容

    Raises:
        ValueError: 不支持的文件类型或解析失败
    """
    ext = os.path.splitext(file_name)[1].lower()

    if ext in IMAGE_EXTENSIONS:
        return f"[图片文件: {file_name}（当前接口为文本模型，无法解析图片内容，请描述图片中的信息）]"

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"不支持的文件类型: {ext}。支持: {', '.join(sorted(SUPPORTED_EXTENSIONS | IMAGE_EXTENSIONS))}")

    try:
        raw = base64.b64decode(file_content)
    except Exception as e:
        raise ValueError(f"Base64 解码失败: {e}")

    size_mb = len(raw) / 1024 / 1024
    if size_mb > 10:
        raise ValueError(f"文件过大: {size_mb:.1f}MB（限制 10MB）")

    try:
        if ext in (".txt", ".csv", ".json", ".md", ".log", ".xml", ".yaml", ".yml", ".toml", ".ini", ".cfg"):
            text = _parse_text(raw)
        elif ext == ".pdf":
            text = _parse_pdf(raw)
        elif ext == ".docx":
            text = _parse_docx(raw)
        elif ext in (".xlsx", ".xls"):
            text = _parse_xlsx(raw)
        else:
            text = _parse_text(raw)
    except Exception as e:
        raise ValueError(f"文件解析失败 ({file_name}): {e}")

    # 限制文本长度
    MAX_CHARS = 50000
    if len(text) > MAX_CHARS:
        logger.warning(f"文件文本超过 {MAX_CHARS} 字符，截断至 {MAX_CHARS}")
        text = text[:MAX_CHARS] + f"\n\n...（文件过长，已截断至 {MAX_CHARS} 字符）"

    return text


# ============================================================
# 统一请求/响应模型
# ============================================================

class ChatRequest(BaseModel):
    message: str = Field(..., description="用户输入", min_length=0, max_length=2000)
    mode: Literal["quick", "smart", "analysis"] = Field("quick", description="对话模式: quick=快速问答 / smart=旧版智能对话（同 quick）/ analysis=深度分析+图表")
    session_id: str = Field("", description="会话 ID（留空自动生成）")
    enable_persistence: bool = Field(False, description="是否启用 SQLite 持久化存储")
    role: Literal["admin", "employee"] = Field("admin", description="角色: admin=管理端（可增删改）, employee=员工端（仅查询）")
    # 文件参数
    file_content: str = Field("", description="文件内容（Base64 编码）")
    file_name: str = Field("", description="原文件名（带扩展名，用于判断文件类型）")
    # 分析参数
    generate_chart: bool = Field(True, description="是否自动生成图表（仅 analysis 模式生效）")


class ApiResponse(BaseModel):
    success: bool
    data: Optional[dict] = None
    error: Optional[str] = None
    session_id: Optional[str] = None
    charts: Optional[list] = None


# ============================================================
# 统一对话端点
# ============================================================

@app.post("/chat", response_model=ApiResponse)
def chat(req: ChatRequest):
    """
    统一对话接口

    通过 mode 参数切换五种模式：
      - single:   单轮对话（三层 Prompt + JSON 约束）
      - tools:    Function Calling 对话（自动调企业 Mock API）
      - multi:    多轮对话（支持上下文裁剪策略）
      - intent:   意图驱动对话（自动分类 + 参数提取 + API 绑定 + 模板输出）
      - analysis: 增强分析对话（意图 + API + 统计分析 + 图表可视化）

    支持文件上传（Base64 编码）：
      将 file_content + file_name 传入，文件文本会自动拼接到 message 后。
    """
    try:
        # 文件解析
        message = req.message
        if req.file_content and req.file_name:
            file_text = parse_file_content(req.file_content, req.file_name)
            message = f"{message}\n\n[上传文件: {req.file_name}]\n{file_text}"
            logger.info(f"[File] 已解析: {req.file_name} ({len(file_text)} chars)")

        session_id, agent = get_or_create_agent(req.session_id, req.enable_persistence)
        # 设置角色权限（admin=管理端可读写，employee=员工端仅查询）
        agent.role = req.role
        logger.info("[Role] session=%s role=%s", session_id, req.role)
        # 将对话模式存入会话元数据，供历史列表显示图标
        if req.enable_persistence:
            try:
                ok = session_store.create_session(session_id, {"mode": req.mode})
                logger.info("[Mode] save session=%s mode=%s ok=%s", session_id, req.mode, ok)
            except Exception as e:
                logger.warning("[Mode] save failed: %s", e)

        if req.mode == "quick":
            result = agent.chat_quick(message)

        elif req.mode == "smart":
            # smart 是旧名称，内部转发到 chat_quick
            result = agent.chat_smart(message)

        elif req.mode == "analysis":
            agent.reset_conversation()
            result = agent.chat_with_analysis(message, generate_chart=req.generate_chart)

        else:
            raise HTTPException(status_code=400, detail=f"未知模式: {req.mode}")

        # 将 reflection 元数据嵌入 data 中传递给前端
        response_data = result.get("data", result)
        if isinstance(response_data, dict):
            if "reflection" in result:
                response_data["reflection"] = result["reflection"]
            if "analysis" in result:
                response_data["analysis"] = result["analysis"]

        return ApiResponse(
            success=result.get("status") == "success",
            data=response_data,
            error=result.get("error"),
            session_id=session_id,
            charts=result.get("charts") or [],
        )
    except HTTPException:
        raise
    except ValueError as e:
        logger.warning(f"[API] 请求参数错误: {e}")
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error(f"[API] /chat 异常: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================
# 其他端点
# ============================================================

@app.get("/api/health")
def health_check():
    """健康检查"""
    return {
        "status": "ok",
        "service": "biz-agent-api",
        "version": "2.0.0",
        "features": {
            "task_tracking": True,
            "session_persistence": True,
            "graded_exceptions": True,
        },
        "active_sessions": len(sessions),
    }


@app.get("/api/sessions")
def list_sessions():
    """查看活跃会话"""
    return {
        "active_sessions": len(sessions),
        "session_ids": list(sessions.keys()),
    }


@app.post("/api/reset/{session_id}")
def reset_session(session_id: str):
    """重置指定会话的对话历史"""
    if session_id in sessions:
        sessions[session_id].reset_conversation()
        logger.info(f"[Session] 重置会话: {session_id}")
        return {"status": "ok", "session_id": session_id}
    raise HTTPException(status_code=404, detail=f"会话 {session_id} 不存在")


@app.post("/api/reset")
def reset_all_sessions():
    """重置所有会话"""
    count = len(sessions)
    for agent in sessions.values():
        agent.reset_conversation()
    TaskTracker.reset_all()
    sessions.clear()
    logger.info("[Session] 重置全部 %d 个会话", count)
    return {"status": "ok", "reset_count": count}


# ============================================================
# 任务追踪 & 会话管理端点（需启用 enable_persistence）
# ============================================================

@app.get("/api/tasks/{session_id}")
def list_session_tasks(session_id: str):
    """列出会话的所有任务"""
    try:
        logs = session_store.get_session_logs(session_id)
        return {"success": True, "data": logs}
    except Exception as e:
        logger.error("[API] 查询任务列表失败: %s", e)
        return {"success": False, "error": str(e)}


@app.get("/api/tasks/{session_id}/{task_id}")
def get_task_detail(session_id: str, task_id: str):
    """获取任务完整执行日志"""
    try:
        logs = session_store.get_task_logs(task_id)
        # 也尝试从内存 TaskTracker 获取
        tracker = TaskTracker._instances.get(session_id)
        memory_logs = tracker.get_task_logs(task_id) if tracker else []
        return {
            "success": True,
            "data": {
                "task_id": task_id,
                "session_id": session_id,
                "db_logs": logs,
                "memory_logs": memory_logs,
            },
        }
    except Exception as e:
        logger.error("[API] 查询任务详情失败: %s", e)
        return {"success": False, "error": str(e)}


@app.get("/api/sessions/{session_id}/messages")
def get_session_messages(session_id: str, limit: int = 100):
    """获取会话消息历史"""
    try:
        messages = session_store.get_messages(session_id, limit=limit)
        return {"success": True, "data": messages}
    except Exception as e:
        logger.error("[API] 查询会话消息失败: %s", e)
        return {"success": False, "error": str(e)}


@app.delete("/api/sessions")
def delete_all_sessions_data():
    """删除所有会话及其数据"""
    count = session_store.delete_all_sessions()
    # 清空内存
    for agent in sessions.values():
        agent.reset_conversation()
    sessions.clear()
    logger.info("[API] 已清除全部 %d 个会话", count)
    return {"success": True, "deleted_count": count}


@app.delete("/api/sessions/{session_id}")
def delete_session_data(session_id: str):
    """删除会话及其所有数据"""
    # 从内存删除
    if session_id in sessions:
        sessions[session_id].reset_conversation()
        del sessions[session_id]
    # 从持久化存储删除
    ok = session_store.delete_session(session_id)
    return {"success": ok, "session_id": session_id, "deleted": ok}


@app.get("/api/logs")
def get_recent_logs(limit: int = 30):
    """获取最近的任务日志"""
    try:
        logs = TaskTracker.read_recent_logs(limit=limit)
        return {"success": True, "data": logs}
    except Exception as e:
        logger.error("[API] 读取日志失败: %s", e)
        return {"success": False, "error": str(e)}


@app.get("/api/stats")
def get_system_stats():
    """获取系统统计信息"""
    total_sessions = len(sessions)
    recent_logs = TaskTracker.read_recent_logs(limit=100)
    total_tasks = len(recent_logs)
    success_tasks = sum(1 for log in recent_logs if log.get("status") == "success")
    return {
        "success": True,
        "data": {
            "active_sessions": total_sessions,
            "recent_tasks": total_tasks,
            "successful_tasks": success_tasks,
            "failed_tasks": total_tasks - success_tasks,
        },
    }


@app.get("/api/history")
def list_history(limit: int = 50):
    """列出持久化的历史会话，含消息预览"""
    try:
        sessions = session_store.list_sessions_with_stats(limit=limit)
        return {"success": True, "data": sessions}
    except Exception as e:
        logger.error("[API] 查询历史失败: %s", e)
        return {"success": False, "error": str(e)}


# ============================================================
# 前端页面
# ============================================================

@app.get("/", response_class=HTMLResponse)
def index():
    """返回商务对话界面"""
    html_path = os.path.join(os.path.dirname(__file__), "static", "index.html")
    with open(html_path, "r", encoding="utf-8") as f:
        return f.read()


# ============================================================
# 启动入口
# ============================================================

def _find_available_port(start: int = 5001, max_attempts: int = 10) -> int:
    """从 start 开始递增查找可用端口，超限后抛出 RuntimeError"""
    import socket
    for port in range(start, start + max_attempts):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            try:
                s.bind(("0.0.0.0", port))
                return port
            except OSError:
                logger.warning("[Port] %d 被占用，尝试下一个...", port)
                continue
    raise RuntimeError(f"无法找到可用端口（尝试范围 {start}-{start + max_attempts - 1}）")


if __name__ == "__main__":
    preferred_port = int(os.getenv("BIZAGENT_PORT", "5001"))
    actual_port = _find_available_port(preferred_port)

    print("=" * 56)
    print("  BizAgent API Server v2.0")
    print("  " + "-" * 40)
    print("  核心功能: 三层 Prompt + Function Calling")
    print("  新增功能: 任务追踪 | 分级异常 | 会话持久化")
    print("  " + "-" * 40)
    print(f"  http://localhost:{actual_port}      (Web 界面)")
    print(f"  http://localhost:{actual_port}/docs (Swagger)")
    print(f"  http://localhost:{actual_port}/api/health")
    if actual_port != preferred_port:
        print(f"  (端口 {preferred_port} 被占用，自动切换至 {actual_port})")
    print("=" * 56)
    uvicorn.run(app, host="0.0.0.0", port=actual_port, log_level="info")
